import customtkinter as ctk
import flet as ft
import os
import pdfplumber
import pandas as pd
import re
import tkinter as tk
from tkinter  import filedialog, messagebox
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, NamedStyle
from openpyxl.utils import get_column_letter
import platform
import subprocess
import unicodedata
import json

@ft.control
class GUI:
    def __init__(self, page: ft.Page):
           
        self.page = page
        self.ruta_origen = []
        self.ruta_destino = ""
        self.page.theme_mode = ft.ThemeMode.LIGHT
        self.contenedor_lista=ft.Column()
        self.visor = ft.Button(
                    "Abrir reporte en Excel",
                    disabled = True,
                    icon=ft.Icons.LOUPE,
                    on_click= self.abrir_archivo_excel
                        )
        self.select_bills = ft.Button("Elige la factura", 
                                icon=ft.Icons.FILE_COPY,
                                on_click=self.seleccionar_archivo)
        self.select_excel = ft.Button("Elige el archivo excel de destino",
                                icon=ft.Icons.LIST, 
                              on_click=self.seleccionar_destino)
        self.procces_bills = ft.Button("Procesar Facturas",
                                    style=ft.ButtonStyle(
                                    color=ft.Colors.WHITE,
                                    bgcolor=ft.Colors.GREEN_800,
                                    overlay_color=ft.Colors.GREEN_600,
                               ),
                               icon=ft.Icons.FILE_DOWNLOAD,
                               icon_color=ft.Colors.WHITE,
                               
                               
                               on_click=self.procesar_todo)
        self.row_procces_bills = ft.Row(
                    controls=[
                       self.select_bills,
                       self.select_excel,
                       self.procces_bills,
                ],
                    alignment=ft.MainAxisAlignment.CENTER, 
                    spacing= 15,
                    
            )
        self.column_visor_excel_and_bills = ft.Column( 
                controls=[
                    self.visor,
                    ft.Text("Facturas seleccionadas"),
                    ft.Container(
                    content=self.contenedor_lista,
                    bgcolor=ft.Colors.BLUE_GREY_600,
                    padding=20,
                    border_radius=10,
                    width=600,
                    height=300
            ),
                    

                ],
                
                spacing= 15
            )
        
         
        self.input_supplier_name = ft.TextField(label="Nombre del proveedor")
        self.alias_supplier_input = ft.TextField(label="Alias (separados por comas)")
        self.cif_number = ft.TextField(label="Escribe el CIF")
        self.list_supplier_view = ft.ListView(height=200, expand=True, spacing=10)
        self.button_save = ft.Button("Añadir Proveedor", on_click=self.save_click)
        
        self.procces_set_bills = ft.Button("Procesar lote de facturas",
                                           style =ft.ButtonStyle(
                                               color=ft.Colors.WHITE,
                                               bgcolor=ft.Colors.GREEN_800,
                                               overlay_color = ft.Colors.GREEN_600,
                                           ),
                                           icon= ft.Icons.FILE_DOWNLOAD,
                                           icon_color= ft.Colors.WHITE,
                                           on_click = self.wrapper_set_bills
                                           )
        self.proccess_set_bills_row = ft.Row(
            controls=[
                self.select_bills,
                self.select_excel,
                self.procces_set_bills
            ],
            alignment = ft.MainAxisAlignment.CENTER,
            spacing= 15,
        )

        self.container_set_bills = ft.Container(
            content = ft.Column(
                controls=[
                    self.proccess_set_bills_row,
                    self.column_visor_excel_and_bills
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            )
        )  
        
        self.scrapper_bills = ft.Container(
            content=ft.Column(
                controls=[
                    self.row_procces_bills,
                    self.column_visor_excel_and_bills
                ],
                alignment=ft.MainAxisAlignment.CENTER,
                horizontal_alignment=ft.CrossAxisAlignment.CENTER,
            ),
            # Aquí puedes añadir otras propiedades como padding, border, etc.
            padding=20,
            border_radius=10
        ) 
        self.button_supplier_visor_row = ft.Button(content="")
        self.button_supplier_delete_row = ft.IconButton(icon=ft.Icons.DELETE, icon_color=ft.Colors.RED, icon_size=30, tooltip="Eliminar")
        self.row_visor_supplier = ft.Row(
            controls = [
                self.button_supplier_visor_row,
                self.button_supplier_delete_row
            ]
        )
        self.container_visor_suppliers = ft.Column(
                controls=[
                    
                ]
                
            )
        

        self.settings_container = ft.Container(

                    content=ft.Column(
                        controls=[
                            ft.Text("Settings", size=20)
                        ]
                    )
                )
        self.supplier_box =ft.Container(
            content=ft.Column(
                [
                    ft.Text("Gestión de proovedores", size=20),
                    self.input_supplier_name,
                    self.alias_supplier_input,
                    self.cif_number,
                    self.list_supplier_view,
                    self.button_save
                ]
            ),bgcolor=ft.Colors.WHITE, # Fondo blanco para contrastar
            padding=20,
            width=350
            )
        
        self.supplier_container = ft.Container(
            content= ft.Row(
                controls=[
                    self.supplier_box,
                    self.container_visor_suppliers
                ]
            ),
            padding=20,
            border=ft.Border.all(1, ft.Colors.BLUE_GREY),
            border_radius=10,
            expand=True   
        )

        
        
        self.settings_tabs = ft.Tab(label="Configuración", icon=ft.Icons.SETTINGS)
        self.suppliers_tabs = ft.Tab(label="Proveedores", icon=ft.Icons.SHOP)
        self.scrapper_bills_tabs = ft.Tab(label="Pasar factura a excel", icon=ft.Icons.FILE_COPY_OUTLINED)
        self.set_bills_tabs = ft.Tab(label="Procesar lote de facturas", icon=ft.Icons.FOLDER_COPY_OUTLINED)

        self.tabs_dic = [
            self.scrapper_bills_tabs,
            self.set_bills_tabs,
            self.suppliers_tabs,
            self.settings_tabs,
            
        ]

        self.tabBar = ft.TabBar(
            tabs=[
                self.scrapper_bills_tabs,
                self.set_bills_tabs,
                self.suppliers_tabs,
                self.settings_tabs,
                
            ]

        )
        self.tabBarView = ft.TabBarView(
            expand=True,
            controls=[
                self.scrapper_bills,
                self.container_set_bills,
                self.supplier_container,
                self.settings_container
            ]
        )
        self.column_tabs = ft.Column(
            expand=True,
            controls = [
                self.tabBar,
                self.tabBarView
            ]
        )

        
        self.tabs = ft.Tabs(
            selected_index=1,
            length=len(self.tabs_dic),
            expand=True,
            content = self.column_tabs,
            on_change= self.on_tab_change 
        )  
    
        self.layout = ft.Row( 
            controls = [
                self.tabs,
            ],expand =True
            )
        
        self.page.add(self.layout)
        
        self.page.update()

    def toggle_suppliers_view(self, e):
        self.supplier_box.visible = True
        self.page.update()

    def abrir_archivo_excel(self):
        """Abre el archivo en el programa predeterminado del SO."""
        try:
            if platform.system() == 'Windows':
                os.startfile(self.ruta_destino)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.call(['open', self.ruta_destino])
            else:  # Linux
                subprocess.call(['xdg-open', self.ruta_destino])
        except Exception as e:
            print(f"No se pudo abrir el archivo: {e}")

    def actualizar_lista_visual(self):
        # Limpiamos y recreamos la lista de textos
        list_names = []
        for path in self.ruta_origen:
            file_name= os.path.basename(path)
            list_names.append(file_name)

        self.contenedor_lista.controls = [
            ft.Text(f"• {names}", size=12, color=ft.Colors.WHITE ) for names in list_names 
        ]
        self.page.update()

    def seleccionar_archivo(self):
        
        # 1. Creamos una ventana raíz oculta de Tkinter
        root = tk.Tk()
        root.withdraw()
        
        # 2. Hacemos que esta ventana esté siempre encima (esto ayuda con el foco)
        root.attributes("-topmost", True)
        
        # 3. Abrimos el diálogo
        archivo_pdf = filedialog.askopenfilenames(
            parent=root, # Le decimos explícitamente que el padre es esta ventana
            title="Selecciona un archivo PDF",
            filetypes=[("Archivos PDF", "*.pdf")]
        )
        
        # 4. Destruimos la ventana raíz para limpiar memoria
        root.destroy()
        
        if archivo_pdf:
            for path in archivo_pdf:
                if path not in self.ruta_origen :
                    self.ruta_origen.append(path)
        self.actualizar_lista_visual()
                    
                    
                    # Si necesitas actualizar la UI de Flet inmediatamente:
                    # self.update()

    def seleccionar_destino(self):
        root = tk.Tk()
        root.withdraw()
        
        # 2. Hacemos que esta ventana esté siempre encima (esto ayuda con el foco)
        root.attributes("-topmost", True)
        self.ruta_destino = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
        print(f"Destino: {self.ruta_destino}")
         # 4. Destruimos la ventana raíz para limpiar memoria
        root.destroy()
    
    def limpiar_concepto(self, text):

        match_date_and_code = re.search(r'Ref Reserva:\s(?P<codigo>\d{8,})\s\((?P<fecha_inicio>\d{1,2}/\d{1,2}/\d{2,4})\s-\s(?P<fecha_fin>\d{1,2}/\d{1,2}/\d{2,4})\)')

    def extraer_texto_completo(self, set_bills):
        
        total_text = ""
        for bill in set_bills:
            with pdfplumber.open(bill) as pdf:
                for page in pdf.pages:
                    # Extrae el texto plano de toda la página
                    total_text += page.extract_text() + "\n"
        
        return total_text
    
    def normalizar_texto(self, total_text):
        # 1. Quitar tildes (convierte 'ó' en 'o')
        total_text = unicodedata.normalize('NFKD', total_text).encode('ascii', 'ignore').decode('utf-8')
        # 2. Convertir a minúsculas
        # 3. Sustituir signos de puntuación extraños por espacios
        total_text = re.sub(r'[,.:;]', ' ', total_text)
        return total_text.lower()
    
    def get_json (self, json_path):
        with open(json_path, 'r', encoding='utf-8')as f:
            data = json.load(f)

        

        all_cifs = []
        all_names = []
        suppliers = data["suppliers"]

        for supplier in suppliers:
            # Aquí 'supplier' es {'name': 'Endesa', 'CIF': 'B12345678', ...}
            # Esto permite acceder por clave sin problemas
            name = supplier.get('name', '')
            cif = supplier.get('CIF', '')
        
        if name:
            all_names.append(re.escape(name))
        if cif:
            all_cifs.append(re.escape(cif))
            
        for alias in supplier.get('alias', []):
            all_names.append(re.escape(alias))
        
        dic_suppliers = {
            "all_names" : all_names,
            "all_cifs" : all_cifs
        }
        return dic_suppliers     
        

    def json_to_dic (self, json_path="suppliers.json"):
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return data

    def limpiar_y_convertir_total(self, match_text):
        if not match_text:
            return None
        
        # 1. Eliminamos todo lo que no sea un dígito
        # Esto elimina "total", ":", "€", espacios, etc.
        solo_numeros = re.sub(r'\D', '', match_text)
        
        # 2. Si después de limpiar no queda nada (ej. el OCR leyó solo letras), devolvemos None
        if not solo_numeros:
            return None
        
        # 3. Aplicamos la lógica de decimales (asumiendo que los últimos 2 son centavos)
        if len(solo_numeros) > 2:
            entero = solo_numeros[:-2]
            decimal = solo_numeros[-2:]
            valor_formateado = f"{entero}.{decimal}"
        else:
            # Si el número es muy pequeño, lo tratamos como decimal
            valor_formateado = f"0.{solo_numeros.zfill(2)}"

        # 4. Convertimos a float de forma segura
        try:
            return float(valor_formateado)
        except ValueError:
            return None 

    def filter_text_by_words (self, normalized_text, json_path ):
        with open(json_path, 'r', encoding='utf-8')as f:
            data = json.load(f)

        all_cifs = []
        all_names = []
        suppliers = data["suppliers"]

        for supplier in suppliers:
            # Aquí 'supplier' es {'name': 'Endesa', 'CIF': 'B12345678', ...}
            # Esto permite acceder por clave sin problemas
            name = supplier.get('name', '')
            cif = supplier.get('CIF', '')
        
        if name:
            all_names.append(re.escape(name))
        if cif:
            all_cifs.append(re.escape(cif))
            
        for alias in supplier.get('alias', []):
            all_names.append(re.escape(alias)) 
    
        regular_expresion_cif_supplier = r'\b('+'|'.join(all_cifs)+r')\b'
        regular_expresion_names_supplier = r'\b(' + '|'.join(all_names) + r')\b'
            
        clean_text = " ".join(normalized_text.split())
        print(f"DEBUG: Texto a analizar: {clean_text[:2000]}") # Imprime los primeros 500 caracteres
        pattern = {
            "Supplier": regular_expresion_names_supplier,
            "Bill" :r'(?i)serie\s*y\s*n[uú]mero\s[,.:]*?\d{4,}/\d{3,}',
            "Date" : r'(?i)fecha(\s+operaci[oó]n)?\s*[.:,\s]*\d{1,2}/\d{1,2}/\d{2,4}',
            "CIF/NIF": regular_expresion_cif_supplier,
            "Subtotal": r'(?i)(subtotal|base\s*imponible|total)\s*[:.,\s]*\s*([\d\s.,]+)\s*', 
            "Total": r'(?i)total\s*[:.,\s]*\s*(\d+[\s.,]?\d{0,2})',
            "Type_IVA": r'(?i)\d{1,2}\s*%' 
        }

        match_iva = re.search(r'(?i)\bIVA\b', clean_text)
    
        # Dividimos el texto en dos partes
        if match_iva:
            # Texto antes de encontrar "IVA"
            bloque_antes = clean_text[:match_iva.start()]
            # Texto después de encontrar "IVA"
            bloque_despues = clean_text[match_iva.start():]
        else:
            bloque_antes = clean_text
            bloque_despues = ""

        # Regex para extraer solo los números (ej: 759 53)
        # Esta regex busca la palabra clave seguida de números separados por espacios o comas
        regex_dinero = r'(?i)(subtotal|base\s*imponible)\s*[:.,\s]*\s*(\d+[\s.,]?\d{0,2})'

        # Buscamos en el bloque correspondiente
        match_base = re.search(regex_dinero, bloque_antes)
        match_iva_sub = re.search(regex_dinero, bloque_despues)

        # Convertimos a formato float usando la función de limpieza anterior
        base_imponible = self.limpiar_y_convertir_total(match_base.group(2)) if match_base else None
        iva_total = self.limpiar_y_convertir_total(match_iva_sub.group(2)) if match_iva_sub else None

        results = {}
        for key, regular_expresion in pattern.items():
            match = re.search(regular_expresion, clean_text, re.IGNORECASE)
            results[key] = match.group(0) if match else None

        total_float = self.limpiar_y_convertir_total(results["Total"])
        results["Total"] = total_float
        results["Subtotal_base_imponible"] = base_imponible
        results["Subtotal_IVA"] = iva_total


        return results
    
    def wrapper_set_bills (self):
        total_text = self.extraer_texto_completo (self.ruta_origen)
        total_text_normalized = self.normalizar_texto(total_text)
        filtered_text = self.filter_text_by_words (normalized_text=total_text_normalized, json_path='suppliers.json')
        print(filtered_text)
        return filtered_text 

    def show_suppliers(self, data):
        data = self.json_to_dic(json_path="suppliers.json")
        list_suppliers = data["suppliers"]
        
        self.container_visor_suppliers.controls = [
            ft.Row(
                controls=[
                    ft.Button(content= s["name"]),
                    ft.IconButton(
                        icon=ft.Icons.DELETE,
                        icon_color=ft.Colors.RED,
                        on_click=lambda e, cif=s["CIF"]: self.delete_supplier_by_cif(cif)
                    )
                ],
                alignment= ft.MainAxisAlignment.START
            )
            for s in list_suppliers
            
        ]
        self.container_visor_suppliers.update()

    def on_tab_change(self, e):
        selected_tab = self.tabs_dic[e.control.selected_index]
        if selected_tab.label == "Proveedores":
            data = self.json_to_dic
            self.show_suppliers(data)
            self.page.update()      


    def save_click(self):
            name = self.input_supplier_name.value
            cif= self.cif_number.value
            alias = self.alias_supplier_input.value

            new_supplier = {
                "name": name,
                "CIF": cif,
                "alias": [a.strip() for a in alias.split(".")]
            }
        
            self.save_supplier( name_value=name, alias_value= alias, cif_value=cif, file_path='suppliers.json')
            self.list_supplier_view.controls.append(ft.Text(f"🏢 {new_supplier['name']}"))
            self.input_supplier_name.value = ""
            self.alias_supplier_input.value = ""
            self.cif_number.value = ""
            self.refresh_supplier_list()
            self.page.update()

    #refresh pantalla


    def save_supplier (self, name_value, alias_value, cif_value, file_path='suppliers.json'):
        new_supplier = {
                "name": name_value,
                "CIF": cif_value,
                "alias": [a.strip() for a in alias_value.split(",")]
            }
        # 1. Cargamos lo que ya existe
        data = self.load_suppliers(file_path)
        
        # 2. Añadimos el nuevo
        data["suppliers"].append(new_supplier)
        print(f"Guardando en: {os.path.abspath(file_path)}")
        # 3. Guardamos todo de vuelta
        with open(file_path, 'w', encoding='utf-8') as file:
            json.dump(data, file, indent=4, ensure_ascii=False)
            file.flush() # Fuerza el volcado al disco
            os.fsync(file.fileno()) # Fuerza la escritura física
        print("Guardado completado exitosamente.") # Mira si esto aparece en la consola

    def load_suppliers (self, file_path):
        if not os.path.exists(file_path) or os.path.getsize(file_path) == 0:
            return{"suppliers": []}
         
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return json.load(file)
        except json.JSONDecodeError:
            return{"suppliers": []}

    def delete_supplier_by_cif(self, cif_to_delete):
        with open('suppliers.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Filtramos por CIF. Como el CIF es único, solo se borrará ese bloque.
        data["suppliers"] = [s for s in data["suppliers"] if s["CIF"] != cif_to_delete]
        
        with open('suppliers.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        
        self.refresh_supplier_list()


    def refresh_supplier_list(self):
        with open("suppliers.json", 'r', encoding='utf-8') as f:
            data = json.load(f)
        self.show_suppliers(data)
        self.page.update()


    def calcular_diferencia_fechas(self, lista_cadenas):
        resultados = []
        
        for item in lista_cadenas:
            # 1.1 Extraer solo los números de los días (lo que está antes de la primera '/')
            # La regex '\d{2}' busca grupos de dos dígitos. 
            # En '(23/07/2026 - 27/07/2026)', esto encontrará '23' y '27'.
            dias = re.findall(r'(\d{2})/', item)
            
            # 1.2 y 1.3 Convertir a lista de ints
            if len(dias) >= 2:
                dias_int = [int(d) for d in dias]
                
                # 2. Restarlos (27 - 23)
                # Nota: Usamos abs() por si el orden de los días está invertido
                resta = abs(dias_int[1] - dias_int[0])

                nights= resta - 1
                
                # 3. Crear el array final
                resultados.append([nights])
            
        return resultados
    
    def extraer_todas_las_tablas(self, ruta_pdf):
        alias_columnas = {
            "Importe": ["importe", "total", "amount", "price", "precio", "subtotal", "Precio unitario"],
            "Concepto": ["concepto", "concept", "descripcion", "description", "descripción"]
        }
        
        # 1. Extraer todas las filas (igual que antes)
        todas_las_filas = []
        with pdfplumber.open(ruta_pdf) as pdf:
            for pagina in pdf.pages:
                tablas = pagina.extract_tables(table_settings={"vertical_strategy": "text", "horizontal_strategy": "text"})
                for tabla in tablas:
                    idx_map = {}
                    fila_cabecera_idx = -1
                    for i, fila in enumerate(tabla):
                        fila_texto = [str(c).lower() if c else "" for c in fila]
                        if any(any(a in celda for lista in alias_columnas.values() for a in lista) for celda in fila_texto):
                            fila_cabecera_idx = i
                            for col_idx, celda in enumerate(fila_texto):
                                for estandar, alias_list in alias_columnas.items():
                                    if any(a in celda for a in alias_list):
                                        idx_map[col_idx] = estandar
                            break
                    for fila in tabla[fila_cabecera_idx + 1:]:
                        fila_dict = {idx_map[i]: fila[i] for i in idx_map.keys() if i < len(fila) and fila[i]}
                        if fila_dict:
                            todas_las_filas.append(fila_dict)



        
        
         # 2. Pre-calcular las reservas globales (para no buscarlas en cada fila)
        texto_total_pdf = self.extraer_texto_completo(ruta_pdf)
        texto_limpio = " ".join(texto_total_pdf.split())
       
        patron = r'Ref\s*Reserva.*?(\d{4,})\s*\((.*?)\s*-\s*(.*?)\)'
        reservas_encontradas = list(re.finditer(patron, texto_total_pdf, re.IGNORECASE | re.DOTALL))
        
        array_plano = [m.group(0) for m in reservas_encontradas]
        

        # 1. Convertir las fechas dentro del paréntesis en dos elementos de una lista. Ej: '(23/07/2026 - 27/072026)'
        #   1.1 Expresión regular que filtre los guiones, los meses y los años. Ej: '23', '27'
        #   1.2 Convertirlos en una lista : un array de dos numeros dentro de una lista Ej: [['23', '27'], ...]
        #   1.3 Convertirlos en int. [[23, 27]...]
        # 2. Restarlos [[23-27]...]
        # 3. Crear un array de numeros  [[5]...]  
        nights = self.calcular_diferencia_fechas(array_plano)
        

        # 1. Filtrar las fechas eliminando todo lo que está fuera del paréntesis. Ej: Ref Reserva: 31682853 (27/04/2026 - 04/05/2026) -> '27/04/2026 - 04/05/2026'
        # 2. Convertirlas en una lista de arrays: cada array tendría la fecha de inicio y la fecha de salida. Ej: [['27/04/2026', '04/05/2026'],...]
        # 3. Restar las fechas dentro de cada array: Mediante un bucle for se itera los arrays mientras otro itera los elementos dentro de los arrays. En esta última iteración se hacen las operaciones DataFrame
        #   3.1 Primera iteración sobre los arrays de la lista
        #       3.1.1 Segunda iteración dentro del array : restar las fechas
        #       3.1.2 Devuelve la diferencia
        #   3.2 La diferencia se añade a un nuevo array
        # Pre-calcular precios de limpieza en el orden de todas_las_filas
        precios_limpieza = []
        for fila in todas_las_filas:
            concepto = str(fila.get("Concepto", "")).lower()
            if "limpieza" in concepto or "arreglo" in concepto:
                precios_limpieza.append(fila.get("Importe", 50))
            else:
                precios_limpieza.append(0)
        # Mantenemos solo los elementos que NO son cero
        precios = [x for x in precios_limpieza if x != 0]
        
        lista_precios_limpieza = [ float (precio.replace('€', '').replace(',', '.').strip())
                                  for precio in precios
                                  ]
        
        precios_reserva = []
        for fila in todas_las_filas:
            precio = fila.get("Importe", "")
            if precios_limpieza[0] != precio:
                precios_reserva.append(fila.get("Importe"))
        set_reserva = set(precios)
        reserva = [x for x in precios_reserva if x not in set_reserva] 
        
        lista_precios_reserva = []
        for v in reserva:
            try:
                if isinstance(v, (int, float)):
                    number = float(v)
                elif isinstance(v, str):
                    clean= v.replace('€', '').replace(',','.').strip()
                    number = float(clean)
                    
                else:
                    continue
                if number > 0:
                    lista_precios_reserva.append(number)
            except (ValueError, TypeError):
                continue
            

                                  
                  
        
        

        # Resultado: [50, 100, 75]
        datos_finales = []
        
        
        
        
        for i, fila in enumerate(todas_las_filas):
            concepto_raw = str(fila.get("Concepto", ""))
            importe_raw = fila.get("Importe") or None
            
            fila_procesada = {
                "Concepto": array_plano[i] if i < len(array_plano) else None,
                "Limpieza": lista_precios_limpieza[i] if i < len(lista_precios_limpieza) else None,
                "Importe": lista_precios_reserva[i] if i < len(lista_precios_reserva)-1 else None,    
            }
            
            
            
            # Si no fue reserva, comprobamos si es limpieza
            
            datos_finales.append(fila_procesada)
                    
        
                    
                    
                    
        return datos_finales
       
    def actualizar_excel(self, nueva_lista_datos, ruta_excel):
        if not nueva_lista_datos: return

        
        try:
            # 1. Preparar datos nuevos
            columnas = ['Concepto', 'Noches', 'Limpieza', 'Importe']
            df_nuevos = pd.DataFrame(nueva_lista_datos)
            df_nuevos.replace([0, '0', '', '0,0', '0,00', '0 €', 0.0], np.nan, inplace=True)
            df_nuevos.dropna(subset=['Importe'], inplace=True)
            fechas_extraidas = df_nuevos['Concepto'].str.extract(r'\((.*?)\)')[0]
            separadas = fechas_extraidas.str.split(' - ', expand=True)
            f_inicio = pd.to_datetime(separadas[0], format='%d/%m/%Y', errors='coerce')
            f_fin = pd.to_datetime(separadas[1], format='%d/%m/%Y', errors='coerce')
            df_nuevos['Noches'] = (f_fin - f_inicio).dt.days - 1
            df_nuevos['Noches'] = df_nuevos['Noches'].fillna(0).astype(int)
            # 2. Cargar datos existentes de forma segura
            if os.path.exists(ruta_excel):
                try:
                    df_existente = pd.read_excel(ruta_excel)
                    # Forzamos las columnas si el archivo existe pero está vacío/mal formado
                    if df_existente.empty or 'Importe' not in df_existente.columns:
                        df_existente = pd.DataFrame(columns= columnas)
                    else:
                        # Si tiene datos, limpiamos los "TOTAL" y ceros
                        df_existente.replace([0, '0', '', '0,0', '0,00', '0 €', 0.0], np.nan, inplace=True)
                        df_existente.dropna(subset=['Importe'], inplace=True)
                        df_existente = df_existente[~df_existente['Concepto'].astype(str).str.contains('TOTAL', case=False, na=False)]
                        df_existente = df_existente[~df_existente['Concepto'].astype(str).str.contains('IVA', case=False, na=False)]
                        df_existente = df_existente[~df_existente['Concepto'].astype(str).str.contains('SUBTOTAL DEL IVA', case=False, na=False)]
                        df_existente = df_existente[~df_existente['Concepto'].astype(str).str.contains('TOTAL SIN IMPUESTOS', case=False, na=False)]
                        df_existente = df_existente[~df_existente['Concepto'].astype(str).str.contains('TOTAL CON IMPUESTOS', case=False, na=False)]
                except:
                    df_existente = pd.DataFrame(columns=columnas)
            else:
                df_existente = pd.DataFrame(columns=columnas)

            # 3. Concatenar y asegurar tipos numéricos
            df_final = pd.concat([df_existente, df_nuevos], ignore_index=True)
            df_final['Importe'] = pd.to_numeric(df_final['Importe'], errors='coerce').fillna(0)
            df_final = df_final[df_final['Importe'] != 0]
            
            with pd.ExcelWriter(ruta_excel, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='Facturas')
                num_filas = len(df_final) + 1
    
                # Accedemos al libro y a la hoja
                worksheet = writer.sheets['Facturas']
                workbook = writer.book # Necesitamos el workbook para registrar el estilo
                # Estilo: Crear un objeto de fuente
                
                header_font = Font(bold=True, color="FFFFFF", name="Arial", size="15")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                alignment_center = Alignment(horizontal='center', vertical='center')
                header_total = Font(bold=True, color="FFFFFF",name="Arial", size="12")
                format_euro = NamedStyle(name="format_euro", number_format='#,##0.00 "€"', alignment= alignment_center)
                workbook.add_named_style(format_euro) # ¡Esto es vital!
                font_text = Font(name='Arial', size=11, bold=False, color='000000')


                # Iterar sobre todas las columnas con datos
                for col in range(1, 5): # De la columna 1 a la 4 (A a D)
                    col_letter = get_column_letter(col)
                    
                    # Calcular el ancho necesario (el +2 es un pequeño margen de espacio)
                    header_val = worksheet.cell(row=1, column=col).value
                    max_length = len(str(header_val)) if header_val else 0
                    
                    for cell in worksheet[col_letter][1:]:
                        try:
                            cell_value = str(cell.value) if cell.value else ""
                            if len(cell_value) > max_length:
                                max_length = len(cell_value)
                        except:
                            pass
                    
                    adjusted_width = max_length + 10
                    worksheet.column_dimensions[col_letter].width = adjusted_width
                
                for row in worksheet.iter_rows(min_row=2, max_row=num_filas, min_col=1, max_col=4):
                    for cell in row:
                        cell.font = font_text
                        cell.alignment = alignment_center

                worksheet.row_dimensions[1].height = 30   
                # Aplicar a la primera fila (encabezados)
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment_center
                    

                for col in range(3, 5): # Columnas C y D (índice 3 y 4)
                    for cell in worksheet[get_column_letter(col)][1:]: # De la fila 2 en adelante
                        cell.style = format_euro
                        

                
                # 1. Definimos el estilo de negrita una vez
                

                # 2. Definimos las etiquetas y sus fórmulas
                etiquetas = [
                    (f'A{num_filas + 2}', 'TOTAL'),
                    (f'A{num_filas + 3}', 'IVA'),
                    (f'A{num_filas + 4}', 'SUBTOTAL DEL IVA'),
                    (f'A{num_filas + 5}', 'TOTAL SIN IMPUESTOS'),
                    (f'A{num_filas + 6}', 'TOTAL CON IMPUESTOS'),
                ]
                # Aplicar a un rango específico de celdas (ej. de A2 a D{num_filas})
                height_row_total = 25
                for row in range(num_filas + 2, num_filas + 7):
                    worksheet.row_dimensions[row].height = height_row_total
      
                # 3. Escribimos etiquetas con negrita
                for celda, texto in etiquetas:
                    worksheet[celda] = texto
                    worksheet[celda].font = header_total
                    worksheet[celda].fill = header_fill
                    worksheet[celda].alignment = alignment_center
                    

                # 4. Escribimos las fórmulas (Asignación directa)
                worksheet[f'B{num_filas + 2}'] = f'=SUM(B2:B{num_filas})'
                worksheet[f'B{num_filas + 2}'].font = font_text 
                worksheet[f'C{num_filas + 2}'] = f'=SUM(C2:C{num_filas})'
                worksheet[f'D{num_filas + 2}'] = f'=SUM(D2:D{num_filas})'
                worksheet[f'C{num_filas + 3}'] = f'=C{num_filas + 2}*0.21'
                worksheet[f'D{num_filas + 3}'] = f'=D{num_filas + 2}*0.21'
                worksheet[f'D{num_filas + 4}'] = f'=C{num_filas + 3} + D{num_filas + 3}'
                worksheet[f'D{num_filas + 5}'] = f'=C{num_filas + 2} + D{num_filas + 2}'
                worksheet[f'D{num_filas + 6}'] = f'=D{num_filas + 4} + D{num_filas + 5}'

                # 5. Aplicamos formato de moneda (previamente definido como NamedStyle o Font)
                # Si quieres aplicar el formato a las celdas de fórmulas:
                for fila in range(num_filas + 2, num_filas + 7):
                    for col in ['C', 'D']:
                        worksheet[f'{col}{fila}'].style = 'format_euro' # O tu formato definido
                        worksheet[f'{col}{fila}'].font = font_text # O tu formato definido
                            

        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error inesperado: {e}")

    def update_set_bills_excel (self, final_data, path_excel):
        if not final_data: return
        try:
            columns = ['Fecha', 'Nº de factura', 'Proveedor', 'CIF/NIF', 'Concepto', 'Base imponible', 'Tipo de IVA', 'Cuota IVA', 'Total Factura']
            df_new = pd.DataFrame(final_data)
            df_new.replace([0, '0', '', '0,0', '0,00', '0 €', 0.0], np.nan, inplace=True)
            
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error inesperado: {e}")

    def procesar_todo(self):
         
        if not self.ruta_origen or not self.ruta_destino:
            messagebox.showwarning("Error", "Selecciona carpeta y destino primero")
            return
        data_list = []
            
        
        for path in self.ruta_origen:    
            if os.path.exists(path):
                try:
                    # Procesar solo ese archivo directamente
                    data = self.extraer_todas_las_tablas(path)
                    if data:
                        data_list.extend(data)
                        
                except Exception as ex:
                    print(f"Error procesando {path}: {ex}")                    
            else:
                messagebox.showinfo("Error", f"La ruta {path} no existe")            
            
        if data_list:
            self.actualizar_excel(data_list, self.ruta_destino)
        else:
            messagebox.showwarning("Aviso", "No se encontraron datos para procesar.")    
        
        self.visor.disabled = False
        self.visor.update()
        

def main(page: ft.Page):
    # Instanciamos nuestra clase y la añadimos a la página
    page.title = "Scrapper bills"
    page.window.width = 900
    page.window.height = 700
    app = GUI(page)
    

    
if __name__ == "__main__":
    ft.run(main)